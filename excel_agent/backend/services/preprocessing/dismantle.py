"""
Wrapper for examples/dismantle_excel.py
Process complex Excel files with merged cells and multi-level headers
完全照搬examples/dismantle_excel.py的逻辑
"""
import sys
import subprocess
import tempfile
import shutil
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from collections import OrderedDict
import json
import uuid
import os

from backend.utils.logging import setup_logger
from backend.config import settings

logger = setup_logger(__name__)


class ExcelDismantler:
    """Dismantle complex Excel files using examples/dismantle_excel.py approach"""
    
    def __init__(self, openai_api_key: str = None):
        """
        Initialize dismantler
        
        Args:
            openai_api_key: OpenAI API key for LLM analysis
        """
        self.openai_api_key = openai_api_key or settings.openai_api_key
    
    @staticmethod
    def unmerge_and_fill_excel(input_path: str, output_path: str) -> Tuple[str, Dict]:
        """
        Unmerge all cells and fill with values
        
        Args:
            input_path: Input Excel file
            output_path: Output Excel file
            
        Returns:
            Tuple of (output_path, merged_info)
        """
        logger.info(f"Unmerging cells in {input_path}")
        
        wb = openpyxl.load_workbook(input_path, data_only=True)
        merged_info = {}
        
        for ws in wb.worksheets:
            logger.info(f"  Processing sheet: {ws.title}")
            sheet_merged_info = []
            
            # Unmerge cells
            for merged_range in list(ws.merged_cells.ranges):
                min_row, min_col, max_row, max_col = (
                    merged_range.min_row, merged_range.min_col, 
                    merged_range.max_row, merged_range.max_col
                )
                value = ws.cell(row=min_row, column=min_col).value
                
                # Store info for header rows only
                if max_row <= 6:
                    sheet_merged_info.append({
                        "merged_range": str(merged_range),
                        "start_cell": (min_row, min_col),
                        "end_cell": (max_row, max_col),
                        "value": value
                    })
                
                # Unmerge
                ws.unmerge_cells(
                    start_row=min_row, start_column=min_col,
                    end_row=max_row, end_column=max_col
                )
                
                # Fill all cells
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        ws.cell(row=row, column=col, value=value)
            
            merged_info[ws.title] = sheet_merged_info
        
        wb.save(output_path)
        logger.info(f"Saved unmerged file to {output_path}")
        
        return output_path, merged_info
    
    @staticmethod
    def get_excel_data(file_path: str, head: int = 6) -> List[str]:
        """
        Get preview of first N rows for each sheet (for LLM analysis)
        完全照搬examples/dismantle_excel.py的get_excel_data
        
        Args:
            file_path: Excel file path
            head: Number of rows
            
        Returns:
            List of sheet preview strings
        """
        try:
            all_sheets_data = pd.read_excel(file_path, sheet_name=None, header=None)
            prompt_parts = []
            
            for sheet_name, data in all_sheets_data.items():
                data.index = data.index + 1
                excel_col_names = [get_column_letter(i + 1) for i in range(len(data.columns))]
                data.columns = excel_col_names
                
                # Handle newlines
                data = data.map(lambda x: str(x).replace('\n', ' ') if isinstance(x, str) else x)
                
                # Convert to string representation
                sheet_first_rows = data.head(head).to_string(index=True)
                sheet_info = f"Sheet: {sheet_name}\n前 {head} 行数据为：\n\n{sheet_first_rows}\n\n---"
                prompt_parts.append(sheet_info)
            
            return prompt_parts
        except Exception as e:
            logger.error(f"提取prompt错误: {e}", exc_info=True)
            return []
    
    def drop_and_merge_excel(self, excel_info: str, merged_info: Dict) -> str:
        """
        Call LLM to analyze which rows to drop and which are headers
        完全照搬examples/dismantle_excel.py的逻辑，但实际调用LLM
        
        Args:
            excel_info: Sheet preview text
            merged_info: Merged cell information
            
        Returns:
            JSON string with labels and header configuration
        """
        try:
            from openai import OpenAI
            
            client = OpenAI(api_key=self.openai_api_key)
            
            prompt = f"""你是一个Excel表格分析专家。请分析以下Excel文件的前6行数据和合并单元格信息，判断：

1. 哪些行是说明性文本（如"单位：元"、"编制"、"审核"等），需要删除
2. 哪些行是表头（可能有多级表头）

【Excel数据】
{excel_info}

【合并单元格信息】
{json.dumps(merged_info, ensure_ascii=False, indent=2)}

请以JSON格式返回结果，格式如下：
[{{"工作表名称": {{
    "labels": [要删除的行号列表],
    "header": [表头行号列表]
}}}}]

示例：
[{{"Sheet1": {{
    "labels": [1],  # 第1行是"单位：元"，需要删除
    "header": [2, 3]  # 第2-3行是多级表头
}}}}]

请只返回JSON，不要其他解释。
"""
            
            response = client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "你是Excel表格结构分析专家。"},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1
            )
            
            result = response.choices[0].message.content
            
            # Extract JSON from markdown if wrapped
            if '```json' in result:
                result = result.split('```json')[1].split('```')[0].strip()
            elif '```' in result:
                result = result.split('```')[1].split('```')[0].strip()
            
            logger.info(f"LLM分析结果: {result}")
            return result
            
        except Exception as e:
            logger.error(f"LLM分析失败: {e}", exc_info=True)
            # Fallback: return default config
            return json.dumps([{sheet_name: {"labels": [], "header": [1]} for sheet_name in merged_info.keys()}])
    
    @staticmethod
    def drop_rows(input_file: str, output_file: str, labels: List[int], sheet_name: str) -> Optional[str]:
        """
        Drop specified rows from sheet
        完全照搬examples/dismantle_excel.py的drop_rows
        
        Args:
            input_file: Input file
            output_file: Output file
            labels: Row indices to drop
            sheet_name: Sheet name
            
        Returns:
            Output file path or None on error
        """
        try:
            df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
            df = df.drop(labels, axis=0, errors='ignore')
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
            return output_file
        except Exception as e:
            logger.error(f'删除指定文本行报错: {e}', exc_info=True)
            return None
    
    @staticmethod
    def deduplication_header(input_file: str, sheet_name: str, 
                            header_rows: List[int]) -> pd.DataFrame:
        """
        Read Excel with multi-level header and deduplicate
        
        Args:
            input_file: Input file
            sheet_name: Sheet name
            header_rows: Header row indices (0-based)
            
        Returns:
            DataFrame with deduplicated single-level header
        """
        # Adjust header rows (convert from 1-based to 0-based)
        header = [h - 1 for h in header_rows] if header_rows else [0]
        
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=header, dtype=object)
        
        # Deduplicate multi-level headers
        if len(header) > 1:
            new_columns = []
            for col in df.columns:
                if isinstance(col, tuple):
                    # Remove unnamed and deduplicate
                    dedup_col = list(OrderedDict.fromkeys(col))
                    valid_header = '-'.join([
                        str(h) for h in dedup_col 
                        if 'Unnamed' not in str(h) and pd.notna(h)
                    ])
                    new_columns.append(valid_header if valid_header else str(col))
                else:
                    new_columns.append(str(col))
            
            df.columns = new_columns
        
        return df
    
    def process_excel_file(self, input_path: Path, output_dir: Path) -> Dict[str, Any]:
        """
        Complete processing pipeline for one Excel file
        完全照搬examples/dismantle_excel.py中main_unmerge_file的逻辑
        
        Args:
            input_path: Input Excel file path
            output_dir: Output directory for processed sheets
            
        Returns:
            Dict containing processed sheets info
        """
        logger.info(f"Processing Excel file: {input_path.name}")
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Step 1: Unmerge cells
        temp_dir = Path(tempfile.mkdtemp())
        unmerged_file = temp_dir / f"unmerged_{uuid.uuid4()}.xlsx"
        
        try:
            _, merged_info = self.unmerge_and_fill_excel(str(input_path), str(unmerged_file))
            
            # Step 2: Get preview data (first 6 rows)
            sheet_info = self.get_excel_data(file_path=str(unmerged_file))
            excel_info = '\n'.join(sheet_info)
            
            # Step 3: Call LLM to analyze structure
            label_info = self.drop_and_merge_excel(excel_info=excel_info, merged_info=merged_info)
            logger.info(f'LLM原始结果:\n{label_info}')
            
            label_info_json = json.loads(label_info.replace('```json', '').replace('```', ''))
            logger.info(f'JSON转换后处理结果:\n{label_info_json}')
            
            # Step 4: Process each sheet according to LLM analysis
            processed_sheets = {}
            
            for sheet_config in label_info_json:
                for sheet_name, config in sheet_config.items():
                    logger.info(f"  Processing sheet: {sheet_name}")
                    
                    labels = config['labels']  # Rows to drop (1-based)
                    header = config['header']  # Header rows (1-based)
                    
                    # Convert to 0-based indices
                    labels_0based = [x - 1 for x in labels]
                    header_0based = [x - len(labels) - 1 for x in header]
                    
                    # Step 4.1: Drop rows
                    drop_file = temp_dir / f'{sheet_name}_modified_{uuid.uuid4()}.xlsx'
                    self.drop_rows(str(unmerged_file), str(drop_file), labels_0based, sheet_name)
                    
                    # Step 4.2: Process header and get final DataFrame
                    df = self.deduplication_header(str(drop_file), sheet_name, header)
                    
                    # Drop empty rows and columns
                    df = df.dropna(how='all')
                    df = df.loc[:, df.notna().any()]
                    
                    # Step 4.3: Save as separate clean Excel file
                    output_file = output_dir / f"{sheet_name}.xlsx"
                    df.to_excel(output_file, index=False, engine='openpyxl')
                    
                    # Step 4.4: Generate profile and field descriptions
                    from backend.services.preprocessing.profiler import DataFrameProfiler
                    profiler = DataFrameProfiler()
                    profile = profiler.profile(df, sheet_name)
                    
                    processed_sheets[sheet_name] = {
                        'file_path': str(output_file),
                        'df': df,
                        'shape': df.shape,
                        'columns': list(df.columns),
                        'header_rows': header,
                        'dropped_rows': labels,
                        'had_merged_cells': len(merged_info.get(sheet_name, [])) > 0,
                        'profile': profile
                    }
                    
                    logger.info(f"    Saved: {output_file.name} {df.shape}")
                    logger.info(f"    Dropped rows: {labels}, Header rows: {header}")
            
            return {
                'original_file': input_path.name,
                'sheets': processed_sheets,
                'sheet_count': len(processed_sheets),
                'merged_info': merged_info
            }
        
        except Exception as e:
            logger.error(f"Excel处理报错: {e}", exc_info=True)
            return None
        
        finally:
            # Cleanup temp files
            if temp_dir.exists():
                shutil.rmtree(temp_dir)

