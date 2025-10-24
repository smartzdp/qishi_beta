# pip install tabulate openpyxl pandas
import json
import logging
import os
import openpyxl
import uuid

import pandas as pd
from openpyxl.utils import get_column_letter
from collections import OrderedDict

logger = logging.getLogger(f'2brain.{__name__}')


def drop_and_merge_excel(excel_info, merged_info):
    return """
[{"复杂表头":
    {
        "labels": [],
        "header": [1, 2]
    }
}]
    """


def get_excel_data(file_path, head=6):
    try:
        all_sheets_data = pd.read_excel(file_path, sheet_name=None, header=None)
        prompt_parts = []

        for sheet_name, data in all_sheets_data.items():
            data.index = data.index + 1
            excel_col_names = [get_column_letter(i + 1) for i in range(len(data.columns))]
            data.columns = excel_col_names  # 替换默认的 `0, 1, 2...` 为 `A, B, C...`

            # 修复：使用 map 替代 applymap，并处理换行符
            data = data.map(lambda x: str(x).replace('\n', ' ') if isinstance(x, str) else x)

            sheet_first_rows = data.head(head).to_markdown(index=True)
            sheet_info = f"Sheet: {sheet_name}\n前 {head} 行数据为：\n\n{sheet_first_rows}\n\n---"
            prompt_parts.append(sheet_info)
        return prompt_parts
    except Exception as e:
        logger.error(f"提取prompt：\n{e}", exc_info=True)


def unmerge_and_fill_excel(input_path, unmerged_file):
    try:
        logger.info("开始取消所有 Sheet 的合并单元格...")

        # 读取 Excel 文件
        wb = openpyxl.load_workbook(input_path, data_only=True)
        logger.info(f"读取 Excel 文件：{input_path} 完成")

        merged_info = {}  # 用于存储合并单元格信息，按每个 sheet 名称分类

        # 遍历所有 Sheet
        for ws in wb.worksheets:
            logger.info(f"正在处理表单：{ws.title} ...")

            # 收集前7行的合并单元格信息
            sheet_merged_info = []

            # 遍历所有合并单元格
            for merged_range in list(ws.merged_cells.ranges):
                # 获取合并单元格的范围
                min_row, min_col, max_row, max_col = (
                    merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                )
                value = ws.cell(row=min_row, column=min_col).value  # 获取合并单元格的主值
                logger.info(f"发现合并单元格：{merged_range}, 值：{value}")

                # 只收集前6行的合并单元格信息
                if max_row <= 6:
                    sheet_merged_info.append({
                        "merged_range": str(merged_range),
                        "start_cell": (min_row, min_col),
                        "end_cell": (max_row, max_col),
                        "value": value
                    })

                # 取消合并单元格
                ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                logger.info(f"取消合并单元格：({min_row}, {min_col}) 到 ({max_row}, {max_col})")

                # 填充所有拆分后的单元格
                logger.info(f"填充单元格的值")
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        ws.cell(row=row, column=col, value=value)

            # 保存当前表单的合并单元格信息
            merged_info[ws.title] = sheet_merged_info
            logger.info(f"表单 {ws.title} 处理完成")

        # 保存处理后的 Excel 文件
        wb.save(unmerged_file)
        logger.info(f"Excel 处理完成，已保存至：{unmerged_file}")
        return unmerged_file, merged_info
    except Exception as e:
        logger.error(f"Excel 处理报错：\n{e}", exc_info=True)


def deduplication_header(drop_file, output_path, header, sheet_name, writer):
    try:
        # 读取指定的 Sheet 文件
        df = pd.read_excel(drop_file, sheet_name=sheet_name, header=header, dtype=object)
        if len(header) == 1:
            df.columns = [df.columns]
        # 去重并拼接表头
        new_columns = []
        for col in df.columns:
            deduplication_col = list(OrderedDict.fromkeys(col))  # 使用 fromkeys 去重并保持顺序
            valid_header = '-'.join([str(header) for header in deduplication_col if 'Unnamed' not in str(header)])
            new_columns.append(valid_header)

        # 更新 DataFrame 的列名
        df.columns = new_columns

        # 将处理后的 DataFrame 写入 Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        logger.error(f'合并多级表头报错:\n{e}', exc_info=True)


def drop_rows(final_unmerged, drop_file, labels, sheet_name):
    try:
        # 读取指定的 Sheet 文件
        df = pd.read_excel(final_unmerged, sheet_name=sheet_name, header=None)

        # 删除指定的行
        df = df.drop(labels, axis=0, errors='ignore')  # errors='ignore' 确保如果没有这些行也不会报错

        # 将修改后的 DataFrame 写入 Excel
        with pd.ExcelWriter(drop_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        return drop_file
    except Exception as e:
        logger.error(f'删除指定文本行报错:\n{e}', exc_info=True)
        return None


def main_unmerge_file(input_file, output_path):
    try:
        # 解除合并文件
        # unmerged_file = "unmerged_file.xlsx"
        unmerged_file = f"unmerged_file_{uuid.uuid4()}.xlsx"  # 使用唯一文件名
        try:
            _, merged_info = unmerge_and_fill_excel(input_file, unmerged_file)

            sheet_info = get_excel_data(file_path=unmerged_file)  # 获取每个表单前6行数据
            excel_info = '\n'.join(sheet_info)
            lable_info = drop_and_merge_excel(excel_info=excel_info,
                                              merged_info=merged_info)  # 根据前6行数据，让大模型指出应该删除的说明性文本以及应该合并的多级表头
            logger.info(f'原始结果:\n{lable_info}')
            lable_info_json = json.loads(lable_info.replace('```json', '').replace('```', ''))

            print(lable_info_json)
            logger.info(f'JSON转换后处理结果:\n{lable_info_json}')

            # 使用同一个 ExcelWriter 写入所有表单
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 处理每个 sheet 的配置
                for sheet in lable_info_json:
                    for sheet_name, config in sheet.items():
                        labels = config['labels']  # 获取需要删除的行
                        header = config['header']  # 获取表头行号
                        labels = [x - 1 for x in labels]
                        header = [x - len(labels) - 1 for x in header]
                        # drop_file = f'{sheet_name}_modified_file.xlsx'  # 临时删除备注信息文件
                        drop_file = f'{sheet_name}_modified_{uuid.uuid4()}.xlsx'

                        # 删除指定行
                        drop_rows(unmerged_file, drop_file, labels, sheet_name)
                        # 处理表头去重并将结果写入同一个文件
                        deduplication_header(drop_file, output_path, header, sheet_name, writer)
                        if os.path.exists(drop_file):
                            os.remove(drop_file)
                        logger.info(f"处理完毕：{sheet_name} 表单保存至 {output_path}")
        finally:
            # 确保临时文件被清理
            if os.path.exists(unmerged_file):
                os.remove(unmerged_file)
        return output_path
    except Exception as e:
        logger.error(f'拆解复杂表头报错{e}', exc_info=True)
        return None


if __name__ == '__main__':
    input_file = '复杂表头.xlsx'
    # 输出文件路径
    output_path1 = "output_path.xlsx"
    main_unmerge_file(input_file, output_path1)
